'''
@author: Abouelhassan Mohammed
'''
from openpyxl import load_workbook

wb = load_workbook(filename = "Gantt_Med.xlsm")

Problem = wb['Problem']
n=Problem['B1'].value
m=Problem['B2'].value


def T():
    T_sheet = wb['T']
    T=[[0 for _ in range(n)] for _ in range(m)]
    for j in range(m):
        for i in range(n):
            if T_sheet.cell(row=i+2, column=j+2).value==None:
                T[j][i]=0
            else: T[j][i]=T_sheet.cell(row=i+2, column=j+2).value


    return T
#print(T())


def d():
    d_sheet = wb['d']
    d=[0 for _ in range(n)]
    for i in range(n):
        if d_sheet.cell(row=i+2, column=2).value==None:
            d[i]=0
        else: d[i]=d_sheet.cell(row=i+2, column=2).value
    return d
#print(d())


def Alpha():
    Alpha_sheet = wb['Alpha']
    Alpha=[[0 for _ in range(n)] for _ in range(m)]
    for j in range(m):
        for i in range(n):
            if Alpha_sheet.cell(row=i+2, column=j+2).value==None :
                Alpha[j][i]=0
            else: Alpha[j][i]=Alpha_sheet.cell(row=i+2, column=j+2).value
    return Alpha
#print(Alpha())


def P():
    P=[[[0 for _ in range(n)] for _ in range(n)] for _ in range(m)]
    for j in range(m):
        P_sheet = wb['P%s'%(j+1)]
        for i in range(n):
            for k in range(n):
                if P_sheet.cell(row=i+2, column=k+2).value==None:
                    P[j][i][k]=0
                else : P[j][i][k]=P_sheet.cell(row=i+2, column=k+2).value

    return P
#print(P())

