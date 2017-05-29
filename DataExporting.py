'''
@author: Abouelhassan Mohammed
'''
from openpyxl import load_workbook
from openpyxl import Workbook
import InteStochasticOpt
import DataImporting


wb = load_workbook(filename = "Gantt_Med.xlsm")
Problem = wb['Problem']
n=Problem['B1'].value
m=Problem['B2'].value

T=DataImporting.T()
Alpha=DataImporting.Alpha()
d=DataImporting.d()
P=DataImporting.P()
print (T)
print (Alpha)
print (d)
print (P)



Npop=30
I=InteStochasticOpt.ProblemSol(Npop,Alpha,T,P,d)
print("optimal makespan is %s"%I[0][0])
#print("optimal makespan is %s"%I)
A=[]
for i in range(Npop):
    A=A+[I[i][0]]
A.reverse()

ExpoX_Y= Workbook()

def WriteX():
    X_sheet=ExpoX_Y.create_sheet("X")
    for j in range(m):
        for i in range(n):
            X_sheet.cell(row=i+2, column=j+2).value=I[0][1][j][i]

WriteX()

def WriteY():
    for j in range(m):
        Y_sheet = ExpoX_Y.create_sheet('Y%s'%(j+1))
        for i in range(n):
            for k in range(n):
                Y_sheet.cell(row=i+2, column=k+2).value=I[0][2][j][i][k]
WriteY()

ExpoX_Y.save("ExpoX_Y.xlsx")


